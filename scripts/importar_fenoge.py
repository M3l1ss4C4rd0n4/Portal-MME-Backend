#!/usr/bin/env python3
"""
Importar datos Fenoge desde Excel a PostgreSQL.
Crea el esquema fenoge con tablas: comunidades, seguimiento.

Uso:
    python3 /home/admonctrlxm/server/scripts/importar_fenoge.py
"""

import os
import re
import sys
import psycopg2
import openpyxl
from datetime import datetime, date

EXCELS_DIR = "/home/admonctrlxm/portal-direccion-mme/excels/fenoge"

DB_CONFIG = dict(
    host=os.getenv("PGHOST", "localhost"),
    port=os.getenv("PGPORT", "5432"),
    database=os.getenv("PGDATABASE", "portal_energetico"),
    user=os.getenv("PGUSER", "mme_user"),
    password=os.getenv("PGPASSWORD", ""),
)


def _num(val):
    if isinstance(val, (int, float)):
        return val
    return None


def _int(val):
    n = _num(val)
    return int(n) if n is not None else None


def _str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        # "1/04/2025" → día/mes/año
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _extract_contract_from_link(link: str | None) -> str | None:
    """Extrae número de contrato '116231-xxx-2025' del URL de SharePoint."""
    if not link:
        return None
    m = re.search(r"116231-(\d+)-(\d{4})", link)
    if m:
        return f"116231-{m.group(1)}-{m.group(2)}"
    return None


# ─────────────────────────────────────────────────────────────────────────────
def create_schema(conn):
    with conn.cursor() as cur:
        cur.execute("CREATE SCHEMA IF NOT EXISTS fenoge;")

        cur.execute("DROP TABLE IF EXISTS fenoge.seguimiento CASCADE;")
        cur.execute("DROP TABLE IF EXISTS fenoge.comunidades CASCADE;")

        cur.execute("""
            CREATE TABLE fenoge.comunidades (
                id               SERIAL PRIMARY KEY,
                fecha_carga      TIMESTAMP DEFAULT NOW(),
                departamento     TEXT,
                municipio        TEXT,
                comunidad        TEXT,
                latitud          NUMERIC,
                longitud         NUMERIC,
                kwp              NUMERIC,
                beneficiarios    INTEGER,
                valor_kwp        NUMERIC,
                valor_proyecto   NUMERIC,
                fase             TEXT,
                lote             INTEGER,
                contratista      TEXT,
                numero_contrato  TEXT,
                operador_red     TEXT,
                fecha_inicio     DATE,
                fecha_fin        DATE
            );
        """)

        cur.execute("""
            CREATE TABLE fenoge.seguimiento (
                id                              SERIAL PRIMARY KEY,
                fecha_carga                     TIMESTAMP DEFAULT NOW(),
                region                          TEXT,
                numero_contrato                 TEXT,
                dia_actualizacion               DATE,
                mes_no                          INTEGER,
                real_financiero                 NUMERIC,
                programado                      NUMERIC,
                real_acumulado_pesos            NUMERIC,
                programado_acumulado_pesos      NUMERIC,
                avance_real_pct                 NUMERIC,
                avance_programado_pct           NUMERIC,
                avance_real_acumulado_pct       NUMERIC,
                avance_programado_acumulado_pct NUMERIC
            );
        """)
    conn.commit()
    print("✓ Esquema fenoge creado")


def import_comunidades(conn):
    path = f"{EXCELS_DIR}/Comunidades Energeticas_fenoge.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Caracterizacion"]

    # Índices de columna (0-based según cabecera):
    # 0:Departamento, 1:Municipio, 2:Comunidad, 3:Lat, 4:Lng, 5:kWp,
    # 6:Beneficiarios, 7:Valor del kWp, 8:Valor del Proyecto(formula),
    # 9:Fase, 10:Lote, 11:Contratista, 12:Link, 13:N°Contrato,
    # 19:FechaInicio, 20:FechaFin, 22:OperadorRed

    rows_ok = 0
    with conn.cursor() as cur:
        for row in ws.iter_rows(min_row=2, values_only=True):
            dept = _str(row[0])
            if not dept:
                continue

            kwp = _num(row[5])
            val_kwp = _num(row[7])
            # "Valor del Proyecto" suele ser fórmula; calcular manualmente
            val_proyecto = (val_kwp * kwp) if (val_kwp and kwp) else _num(row[8])

            # Número de contrato: puede ser valor literal o fórmula
            contrato_raw = row[13]
            if isinstance(contrato_raw, str) and contrato_raw.startswith("="):
                contrato_raw = _extract_contract_from_link(_str(row[12]))
            numero_contrato = _str(contrato_raw)

            cur.execute("""
                INSERT INTO fenoge.comunidades
                    (departamento, municipio, comunidad, latitud, longitud,
                     kwp, beneficiarios, valor_kwp, valor_proyecto, fase, lote,
                     contratista, numero_contrato, operador_red, fecha_inicio, fecha_fin)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                dept, _str(row[1]), _str(row[2]),
                _num(row[3]), _num(row[4]),
                kwp, _int(row[6]), val_kwp, val_proyecto,
                _str(row[9]), _int(row[10]), _str(row[11]),
                numero_contrato, _str(row[22]),
                _parse_date(row[19]), _parse_date(row[20]),
            ))
            rows_ok += 1

    conn.commit()
    print(f"✓ Comunidades importadas: {rows_ok} registros")


def import_seguimiento(conn):
    path = f"{EXCELS_DIR}/comunidades_seguimiento_fenoge.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Seguimiento"]

    # 0:REGIÓN, 1:NUMERO DE CONTRATO, 2:DÍA ACTUALIZACIÓN, 3:MES No.,
    # 4:REAL, 5:REAL FINANCIERO, 6:PROGRAMADO, 7:REAL ACUMULADO $,
    # 8:PROGRAMADO ACUMULADO $, 9:AVANCE REAL %, 10:AVANCE PROG %,
    # 11:AVANCE REAL ACUMULADO %, 12:AVANCE PROG ACUMULADO %

    rows_ok = 0
    with conn.cursor() as cur:
        for row in ws.iter_rows(min_row=2, values_only=True):
            region = _str(row[0])
            if not region:
                continue

            cur.execute("""
                INSERT INTO fenoge.seguimiento
                    (region, numero_contrato, dia_actualizacion, mes_no,
                     real_financiero, programado,
                     real_acumulado_pesos, programado_acumulado_pesos,
                     avance_real_pct, avance_programado_pct,
                     avance_real_acumulado_pct, avance_programado_acumulado_pct)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                region, _str(row[1]), _parse_date(row[2]), _int(row[3]),
                _num(row[5]), _num(row[6]),
                _num(row[7]), _num(row[8]),
                _num(row[9]),  _num(row[10]),
                _num(row[11]), _num(row[12]),
            ))
            rows_ok += 1

    conn.commit()
    print(f"✓ Seguimiento importado: {rows_ok} registros")


if __name__ == "__main__":
    print("Conectando a PostgreSQL…")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        sys.exit(1)

    try:
        create_schema(conn)
        import_comunidades(conn)
        import_seguimiento(conn)
        print("\n✅ Importación Fenoge completada exitosamente")

        # Verificación rápida
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM fenoge.comunidades")
            n_ce = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM fenoge.seguimiento")
            n_seg = cur.fetchone()[0]
            cur.execute("SELECT SUM(kwp), SUM(beneficiarios), SUM(valor_proyecto) FROM fenoge.comunidades")
            tot = cur.fetchone()
        print(f"\n--- Resumen ---")
        print(f"  CEs importadas   : {n_ce}")
        print(f"  Seguimiento rows : {n_seg}")
        print(f"  Total kWp        : {tot[0]}")
        print(f"  Total benefic.   : {tot[1]}")
        print(f"  Total inversión  : ${tot[2]:,.0f}" if tot[2] else "  Total inversión  : N/A")
    finally:
        conn.close()
